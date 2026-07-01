#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""แปลงไฟล์ GRC_Progress_Input.xlsx -> data.json สำหรับ Dashboard GRC
วิธีใช้:  python convert.py GRC_Progress_Input.xlsx data.json
ต้องติดตั้ง:  pip install openpyxl
"""
import sys, json
from openpyxl import load_workbook

TH_FULL=["มกราคม","กุมภาพันธ์","มีนาคม","เมษายน","พฤษภาคม","มิถุนายน",
         "กรกฎาคม","สิงหาคม","กันยายน","ตุลาคม","พฤศจิกายน","ธันวาคม"]

def num(v):
    try: return float(v) if v not in (None,'') else 0.0
    except: return 0.0
def txt(v):
    return '' if v is None else str(v).replace('\xa0',' ').strip()

def short(name):
    s=name.split(' • ')[0].strip()
    return (s[:38]+'…') if len(s)>40 else s

def convert(xlsx_path, out_path='data.json'):
    wb=load_workbook(xlsx_path, data_only=True)
    s=wb['ตั้งค่า']
    planTitle=txt(s['C4'].value); source=txt(s['C5'].value)
    year=txt(s['C6'].value) or '2569'
    month=txt(s['C7'].value) or 'มิถุนายน'
    curMonth=TH_FULL.index(month) if month in TH_FULL else 5
    updated=f'{month} {year}'

    aw=wb['กิจกรรม']
    TCOL=6; ACOL=18; NCOL=30  # 1-based: F, R, AD
    acts=[]
    for row in range(3, aw.max_row+1):
        idv=aw.cell(row,1).value
        if idv in (None,''): continue
        name=txt(aw.cell(row,2).value).replace('\n• ',' • ').replace('\n',' • ')
        unit=txt(aw.cell(row,3).value)
        w=num(aw.cell(row,4).value)
        out=txt(aw.cell(row,5).value)
        t=[num(aw.cell(row,TCOL+i).value) for i in range(12)]
        a=[num(aw.cell(row,ACOL+i).value) for i in range(12)]
        q=[txt(aw.cell(row,NCOL+i).value) for i in range(4)]
        notes=['']*12
        for qi,idx in zip(range(4),[2,5,8,11]): notes[idx]=q[qi]
        acts.append(dict(id=str(int(idv)),aid='GRC-%02d'%int(idv),name=name,short=short(name),
                         unit=unit,w=w,t=t,a=a,notes=notes,out=out))
    tw=sum(x['w'] for x in acts) or 1
    cumT=[round(sum(x['w']*x['t'][m] for x in acts)/tw,1) for m in range(12)]
    cumA=[round(sum(x['w']*x['a'][m] for x in acts)/tw,1) for m in range(12)]
    data=dict(planTitle=planTitle,source=source,updated=updated,curMonth=curMonth,
              nact=len(acts),cumT=cumT,cumA=cumA,acts=acts)
    with open(out_path,'w',encoding='utf-8') as f:
        json.dump(data,f,ensure_ascii=False,indent=1)
    print(f'✓ เขียน {out_path} สำเร็จ — {len(acts)} กิจกรรม, เดือนปัจจุบัน {month} {year}')
    return data

if __name__=='__main__':
    xlsx = sys.argv[1] if len(sys.argv)>1 else 'GRC_Progress_Input.xlsx'
    outp = sys.argv[2] if len(sys.argv)>2 else 'data.json'
    convert(xlsx, outp)
